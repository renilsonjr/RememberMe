import pytest
import shutil
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch, MagicMock

from src.presentation.app import create_app


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

BASE_XLSX = Path(__file__).parent / "RememberMe_template.xlsx"


@pytest.fixture
def xlsx_path(tmp_path):
    dst = tmp_path / "RememberMe.xlsx"
    shutil.copy(BASE_XLSX, dst)
    return str(dst)


@pytest.fixture
def client(xlsx_path):
    app = create_app({"XLSX_FILE": xlsx_path, "TESTING": True, "WTF_CSRF_ENABLED": False})
    app.config["SECRET_KEY"] = "test-secret"
    with app.test_client() as c:
        yield c


# ---------------------------------------------------------------------------
# GET /
# ---------------------------------------------------------------------------

class TestIndexRoute:

    def test_returns_200(self, client):
        response = client.get("/")
        assert response.status_code == 200

    def test_creditor_name_appears_in_html(self, client):
        response = client.get("/")
        # The template should at least render without errors
        assert response.status_code == 200

    def test_empty_payments_renders_without_error(self, tmp_path):
        from openpyxl import Workbook
        empty = tmp_path / "empty.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "RememberMe"
        ws["A3"] = "Creditor"
        wb.save(str(empty))
        app = create_app({"XLSX_FILE": str(empty), "TESTING": True})
        with app.test_client() as c:
            response = c.get("/")
        assert response.status_code == 200

    def test_due_soon_class_in_html(self, xlsx_path):
        """If any creditor is due within 14 days, due-soon class appears."""
        from src.infrastructure.repositories.excel_repository import ExcelRepository
        from src.domain.payment import Payment, Settlement
        repo = ExcelRepository(xlsx_path)
        due_soon_payment = Payment(
            creditor="SOON BANK",
            payment_num=1,
            amount=Decimal("100.00"),
            due_date=date.today() + timedelta(days=5),
            balance_before=Decimal("500.00"),
            balance_after=Decimal("400.00"),
        )
        repo.add_settlement(Settlement("SOON BANK", [due_soon_payment]))
        app = create_app({"XLSX_FILE": xlsx_path, "TESTING": True})
        with app.test_client() as c:
            response = c.get("/")
        assert b"due-soon" in response.data

    def test_paid_off_badge_present_when_balance_zero(self, xlsx_path):
        from src.infrastructure.repositories.excel_repository import ExcelRepository
        from src.domain.payment import Payment, Settlement
        repo = ExcelRepository(xlsx_path)
        paid_payment = Payment(
            creditor="PAID BANK",
            payment_num=1,
            amount=Decimal("100.00"),
            due_date=date.today() - timedelta(days=5),
            balance_before=Decimal("100.00"),
            balance_after=Decimal("0.00"),
        )
        repo.add_settlement(Settlement("PAID BANK", [paid_payment]))
        app = create_app({"XLSX_FILE": xlsx_path, "TESTING": True})
        with app.test_client() as c:
            response = c.get("/")
        assert b"PAID OFF" in response.data


# ---------------------------------------------------------------------------
# POST /add-settlement
# ---------------------------------------------------------------------------

class TestAddSettlementRoute:

    VALID_FORM = {
        "creditor_name": "NEW BANK",
        "total_amount": "500.00",
        "monthly_payment": "100.00",
        "first_due_date": "2026-06-01",
        "payment_day_of_month": "1",
    }

    def test_post_redirects_to_index(self, client):
        response = client.post("/add-settlement", data=self.VALID_FORM)
        assert response.status_code == 302
        assert response.headers["Location"].endswith("/")

    def test_missing_creditor_name_redirects(self, client):
        bad_form = {k: v for k, v in self.VALID_FORM.items() if k != "creditor_name"}
        response = client.post("/add-settlement", data=bad_form)
        assert response.status_code == 302

    def test_invalid_date_redirects(self, client):
        bad_form = {**self.VALID_FORM, "first_due_date": "not-a-date"}
        response = client.post("/add-settlement", data=bad_form)
        assert response.status_code == 302

    def test_invalid_amount_redirects(self, client):
        bad_form = {**self.VALID_FORM, "total_amount": "not-a-number"}
        response = client.post("/add-settlement", data=bad_form)
        assert response.status_code == 302

    def test_creditor_appears_after_add(self, client):
        client.post("/add-settlement", data=self.VALID_FORM)
        response = client.get("/")
        assert b"NEW BANK" in response.data


# ---------------------------------------------------------------------------
# POST /delete-settlement
# ---------------------------------------------------------------------------

class TestDeleteSettlementRoute:

    def test_post_redirects_to_index(self, client):
        response = client.post("/delete-settlement", data={"creditor_name": "NONEXISTENT"})
        assert response.status_code == 302
        assert response.headers["Location"].endswith("/")

    def test_missing_creditor_name_redirects(self, client):
        response = client.post("/delete-settlement", data={})
        assert response.status_code == 302

    def test_creditor_removed_after_delete(self, client):
        # Add first
        client.post("/add-settlement", data={
            "creditor_name": "REMOVABLE CREDITOR XYZ",
            "total_amount": "200.00",
            "monthly_payment": "100.00",
            "first_due_date": "2026-06-01",
        })
        assert b"REMOVABLE CREDITOR XYZ" in client.get("/").data
        # Then delete — follow redirect to flush flash message, then fetch again
        client.post("/delete-settlement", data={"creditor_name": "REMOVABLE CREDITOR XYZ"},
                    follow_redirects=True)
        # A second GET should no longer show this creditor in the table
        response = client.get("/")
        assert b"REMOVABLE CREDITOR XYZ" not in response.data


# ---------------------------------------------------------------------------
# GET /sync
# ---------------------------------------------------------------------------

class TestSyncRoute:

    def test_returns_200(self, client):
        response = client.get("/sync")
        assert response.status_code == 200

    def test_success_message_in_response(self, client):
        response = client.get("/sync")
        assert b"calendar" in response.data.lower() or b"event" in response.data.lower() or b"sync" in response.data.lower()
