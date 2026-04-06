import pytest
import shutil
from datetime import date
from pathlib import Path

from excel_writer import generate_payment_rows, append_to_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

BASE_XLSX = Path(__file__).parent / "RememberMe_template.xlsx"


def _make_xlsx(tmp_path):
    """Copy the real template into tmp_path so append_to_excel has a base file."""
    dst = tmp_path / "RememberMe.xlsx"
    shutil.copy(BASE_XLSX, dst)
    return str(dst)


# ---------------------------------------------------------------------------
# 1. generate_payment_rows
# ---------------------------------------------------------------------------

class TestGeneratePaymentRows:

    def test_returns_list(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert isinstance(rows, list)

    def test_returns_dicts(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert all(isinstance(r, dict) for r in rows)

    def test_dict_has_reader_contract_keys(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        required = {"creditor", "payment_num", "amount", "due_date",
                    "balance_before", "balance_after"}
        for row in rows:
            assert required.issubset(row.keys())

    def test_correct_number_of_rows_exact_division(self):
        # 300 / 100 = 3 payments exactly
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert len(rows) == 3

    def test_correct_number_of_rows_with_remainder(self):
        # 350 / 100 = 3 full + 1 partial = 4 payments
        rows = generate_payment_rows("Test", 350.0, 100.0, date(2026, 5, 1))
        assert len(rows) == 4

    def test_single_payment_when_total_less_than_monthly(self):
        rows = generate_payment_rows("Test", 50.0, 100.0, date(2026, 5, 1))
        assert len(rows) == 1
        assert rows[0]["amount"] == pytest.approx(50.0)

    def test_payment_num_sequential_from_one(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert [r["payment_num"] for r in rows] == [1, 2, 3]

    def test_creditor_set_on_all_rows(self):
        rows = generate_payment_rows("ACME", 300.0, 100.0, date(2026, 5, 1))
        assert all(r["creditor"] == "ACME" for r in rows)

    def test_first_due_date_used_for_payment_1(self):
        first = date(2026, 5, 15)
        rows = generate_payment_rows("Test", 300.0, 100.0, first)
        assert rows[0]["due_date"] == first

    def test_interval_dates_when_no_payment_day_given(self):
        """Without payment_day_of_month, dates advance by 28 days."""
        first = date(2026, 5, 1)
        rows = generate_payment_rows("Test", 300.0, 100.0, first)
        assert rows[1]["due_date"] == date(2026, 5, 29)
        assert rows[2]["due_date"] == date(2026, 6, 26)

    def test_monthly_dates_use_payment_day_of_month(self):
        """With payment_day_of_month=15, dates fall on the 15th each month."""
        first = date(2026, 5, 15)
        rows = generate_payment_rows("Test", 300.0, 100.0, first,
                                     payment_day_of_month=15)
        assert rows[0]["due_date"] == date(2026, 5, 15)
        assert rows[1]["due_date"] == date(2026, 6, 15)
        assert rows[2]["due_date"] == date(2026, 7, 15)

    def test_date_clamped_to_last_day_of_short_month(self):
        """Day 31 in February should clamp to Feb 28 (or 29 in leap year)."""
        first = date(2026, 1, 31)
        rows = generate_payment_rows("Test", 300.0, 100.0, first,
                                     payment_day_of_month=31)
        assert rows[1]["due_date"] == date(2026, 2, 28)
        assert rows[2]["due_date"] == date(2026, 3, 31)

    def test_full_payments_have_standard_amount(self):
        rows = generate_payment_rows("Test", 350.0, 100.0, date(2026, 5, 1))
        # First 3 payments are full
        for row in rows[:3]:
            assert row["amount"] == pytest.approx(100.0)

    def test_last_payment_amount_is_remainder(self):
        # 350 - 3*100 = 50
        rows = generate_payment_rows("Test", 350.0, 100.0, date(2026, 5, 1))
        assert rows[-1]["amount"] == pytest.approx(50.0)

    def test_last_payment_sum_equals_total(self):
        total = 347.83
        rows = generate_payment_rows("Test", total, 100.0, date(2026, 5, 1))
        assert sum(r["amount"] for r in rows) == pytest.approx(total, abs=0.01)

    def test_balance_before_first_row_equals_total(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert rows[0]["balance_before"] == pytest.approx(300.0)

    def test_balance_before_chains_correctly(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert rows[1]["balance_before"] == pytest.approx(200.0)
        assert rows[2]["balance_before"] == pytest.approx(100.0)

    def test_balance_after_is_none(self):
        """balance_after is always None — Excel formula handles it."""
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert all(r["balance_after"] is None for r in rows)

    def test_notes_empty_for_non_final_rows(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        for row in rows[:-1]:
            assert row.get("notes", "") == ""

    def test_notes_settlement_complete_for_final_row(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert "Settlement" in rows[-1].get("notes", "")

    def test_exact_division_no_extra_row(self):
        rows = generate_payment_rows("Test", 300.0, 100.0, date(2026, 5, 1))
        assert len(rows) == 3
        assert rows[-1]["amount"] == pytest.approx(100.0)


# ---------------------------------------------------------------------------
# 2. append_to_excel
# ---------------------------------------------------------------------------

class TestAppendToExcel:

    def test_rows_appended_after_existing_data(self, tmp_path):
        from reader import load_payments
        filepath = _make_xlsx(tmp_path)
        before = len(load_payments(filepath))
        new_rows = generate_payment_rows("NEW CO", 200.0, 100.0, date(2026, 6, 1))
        append_to_excel(new_rows, filepath)
        after = load_payments(filepath)
        assert len(after) == before + len(new_rows)

    def test_appended_creditor_name_present(self, tmp_path):
        from reader import load_payments
        filepath = _make_xlsx(tmp_path)
        new_rows = generate_payment_rows("ZENITH BANK", 200.0, 100.0, date(2026, 6, 1))
        append_to_excel(new_rows, filepath)
        creditors = [p["creditor"] for p in load_payments(filepath)]
        assert "ZENITH BANK" in creditors

    def test_balance_after_formula_readable_by_load_payments(self, tmp_path):
        """After append, load_payments must compute balance_after correctly."""
        from reader import load_payments
        filepath = _make_xlsx(tmp_path)
        new_rows = generate_payment_rows("FORMULA CO", 200.0, 100.0, date(2026, 6, 1))
        append_to_excel(new_rows, filepath)
        payments = load_payments(filepath)
        formula_co = [p for p in payments if p["creditor"] == "FORMULA CO"]
        assert formula_co[0]["balance_after"] == pytest.approx(100.0)
        assert formula_co[1]["balance_after"] == pytest.approx(0.0)

    def test_blank_separator_row_between_creditors(self, tmp_path):
        """Two appends should leave a blank row between the creditor groups."""
        import openpyxl
        filepath = _make_xlsx(tmp_path)
        rows_a = generate_payment_rows("ALPHA", 100.0, 100.0, date(2026, 6, 1))
        rows_b = generate_payment_rows("BETA", 100.0, 100.0, date(2026, 7, 1))
        append_to_excel(rows_a, filepath)
        append_to_excel(rows_b, filepath)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        # There must be at least one None between ALPHA and BETA
        alpha_last = max(i for i, v in enumerate(values) if v == "ALPHA")
        beta_first = min(i for i, v in enumerate(values) if v == "BETA")
        assert beta_first > alpha_last + 1  # gap between them

    def test_raises_if_file_not_found(self, tmp_path):
        new_rows = generate_payment_rows("Test", 100.0, 100.0, date(2026, 6, 1))
        with pytest.raises(FileNotFoundError):
            append_to_excel(new_rows, str(tmp_path / "nonexistent.xlsx"))

    def test_round_trip_amounts(self, tmp_path):
        from reader import load_payments
        filepath = _make_xlsx(tmp_path)
        new_rows = generate_payment_rows("ROUND TRIP", 350.0, 100.0, date(2026, 6, 1))
        append_to_excel(new_rows, filepath)
        payments = load_payments(filepath)
        rt = [p for p in payments if p["creditor"] == "ROUND TRIP"]
        assert len(rt) == 4
        assert rt[0]["amount"] == pytest.approx(100.0)
        assert rt[-1]["amount"] == pytest.approx(50.0)
