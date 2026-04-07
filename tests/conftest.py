"""Shared test fixtures and configuration."""

import os
import tempfile
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import Workbook

from src.domain.payment import Payment, Settlement
from src.infrastructure.repositories.excel_repository import ExcelRepository


@pytest.fixture
def temp_excel_file():
    """Create a temporary Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        temp_path = f.name

    # Create a workbook with headers
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "RememberMe"
    ws["A3"] = "Creditor"
    ws["B3"] = "Payment #"
    ws["C3"] = "Amount"
    ws["D3"] = "Due Date"
    ws["E3"] = "Balance Before"
    ws["F3"] = "Balance After"

    wb.save(temp_path)

    yield temp_path

    # Cleanup
    if os.path.exists(temp_path):
        os.unlink(temp_path)


@pytest.fixture
def excel_repository(temp_excel_file):
    """Create an Excel repository with a temporary file."""
    return ExcelRepository(temp_excel_file)


@pytest.fixture
def sample_payment():
    """Create a sample payment for testing."""
    return Payment(
        creditor="TEST CREDITOR",
        payment_num=1,
        amount=Decimal("100.00"),
        due_date=date.today(),
        balance_before=Decimal("500.00"),
        balance_after=Decimal("400.00"),
    )


@pytest.fixture
def sample_settlement(sample_payment):
    """Create a sample settlement for testing."""
    payments = [
        sample_payment,
        Payment(
            creditor="TEST CREDITOR",
            payment_num=2,
            amount=Decimal("100.00"),
            due_date=date(2026, 5, 6),
            balance_before=Decimal("400.00"),
            balance_after=Decimal("300.00"),
        ),
        Payment(
            creditor="TEST CREDITOR",
            payment_num=3,
            amount=Decimal("100.00"),
            due_date=date(2026, 6, 6),
            balance_before=Decimal("300.00"),
            balance_after=Decimal("200.00"),
        ),
    ]
    return Settlement("TEST CREDITOR", payments)
