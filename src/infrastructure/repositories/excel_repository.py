"""Excel-based implementation of PaymentRepository."""

import os
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.domain.payment import Payment, Settlement
from src.domain.errors import SettlementNotFoundError, InvalidPaymentError


# Cell fill / font colours matching the existing RememberMe template
_FILL_BLUE = PatternFill("solid", fgColor="FFDDEEFF")   # input columns A-E
_FILL_GREY = PatternFill("solid", fgColor="FFF2F2F2")   # formula columns F-I
_FILL_GREEN = PatternFill("solid", fgColor="FFE2EFDA")  # final-row notes col J
_FONT_BLUE = Font(color="FF0000FF")
_FONT_BLACK = Font(color="FF000000")


class ExcelRepository:
    """Payment data repository backed by Excel file."""

    def __init__(self, filepath: str):
        """Initialize with path to Excel file."""
        self.filepath = filepath
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"No such file: {filepath}")

    def get_all_payments(self) -> list[Payment]:
        """Load all payments from Excel."""
        try:
            wb = openpyxl.load_workbook(self.filepath, data_only=True)
        except Exception as exc:
            raise IOError(f"Failed to load Excel file: {exc}")

        ws = wb.active
        payments = []

        for row in ws.iter_rows(min_row=4, values_only=True):
            payment = self._row_to_payment(row)
            if payment:
                payments.append(payment)

        return payments

    def get_all_settlements(self) -> list[Settlement]:
        """Get all settlements (grouped payments by creditor)."""
        payments = self.get_all_payments()

        # Group payments by creditor, preserving order
        settlements_dict: dict[str, list[Payment]] = {}
        for payment in payments:
            if payment.creditor not in settlements_dict:
                settlements_dict[payment.creditor] = []
            settlements_dict[payment.creditor].append(payment)

        return [
            Settlement(creditor, payments)
            for creditor, payments in settlements_dict.items()
        ]

    def get_settlement(self, creditor: str) -> Settlement | None:
        """Get a single settlement by creditor name."""
        try:
            settlements = self.get_all_settlements()
            for settlement in settlements:
                if settlement.creditor == creditor:
                    return settlement
            return None
        except Exception as exc:
            raise IOError(f"Failed to retrieve settlement: {exc}")

    def add_settlement(self, settlement: Settlement) -> None:
        """Add a new settlement to Excel file."""
        # Convert Settlement to payment rows
        payment_rows = [
            {
                "creditor": p.creditor,
                "payment_num": p.payment_num,
                "amount": float(p.amount),
                "due_date": p.due_date,
                "balance_before": float(p.balance_before),
                "balance_after": None,  # Excel formula
                "notes": "← Settlement complete!" if p.is_final() else "",
            }
            for p in settlement.payments
        ]

        self._append_to_excel(payment_rows)

    def remove_settlement(self, creditor: str) -> None:
        """Remove a settlement from Excel file."""
        self._remove_creditor(creditor)

    def update_settlement(self, settlement: Settlement) -> None:
        """Update an existing settlement (remove and re-add)."""
        self.remove_settlement(settlement.creditor)
        self.add_settlement(settlement)

    @staticmethod
    def _row_to_payment(row: tuple) -> Payment | None:
        """Convert Excel row to Payment entity."""
        if not row or len(row) < 6:
            return None

        creditor, payment_num, amount, due_date, balance_before, balance_after = row[:6]

        # Skip blank separator rows
        if creditor is None:
            return None

        # Normalize due_date (may come as datetime from openpyxl)
        if isinstance(due_date, datetime):
            due_date = due_date.date()

        # Compute balance_after if it's a formula (None)
        if balance_after is None:
            balance_after = (balance_before or 0.0) - (amount or 0.0)

        try:
            return Payment(
                creditor=str(creditor),
                payment_num=int(payment_num) if payment_num is not None else 0,
                amount=Decimal(str(amount)) if amount is not None else Decimal("0"),
                due_date=due_date,
                balance_before=Decimal(str(balance_before)) if balance_before is not None else Decimal("0"),
                balance_after=Decimal(str(balance_after)),
            )
        except (ValueError, TypeError) as exc:
            raise InvalidPaymentError(f"Invalid payment data: {exc}")

    def _append_to_excel(self, payments: list[dict]) -> None:
        """Append a list of payment dicts to Excel file."""
        wb = openpyxl.load_workbook(self.filepath, data_only=False)
        ws = wb.active

        # Find the first empty row after all existing data
        last_data_row = 3  # at minimum, 3 header rows exist
        for row_idx in range(ws.max_row, 3, -1):
            if ws.cell(row=row_idx, column=1).value is not None:
                last_data_row = row_idx
                break

        # Insert one blank separator row
        separator_row = last_data_row + 1
        first_data_row = separator_row + 1

        for i, payment in enumerate(payments):
            excel_row = first_data_row + i

            # Input columns (A-E): blue fill, blue font
            def write_input(col, value):
                cell = ws.cell(row=excel_row, column=col, value=value)
                cell.fill = _FILL_BLUE
                cell.font = _FONT_BLUE

            write_input(1, payment["creditor"])
            write_input(2, payment["payment_num"])
            write_input(3, payment["amount"])
            write_input(4, payment["due_date"])
            write_input(5, payment["balance_before"])

            # Formula columns (F-I): grey fill, black font
            def write_formula(col, formula):
                cell = ws.cell(row=excel_row, column=col, value=formula)
                cell.fill = _FILL_GREY
                cell.font = _FONT_BLACK

            r = excel_row
            write_formula(6, f"=E{r}-C{r}")
            write_formula(7, f'=IF(F{r}=0,"✅ PAID OFF",IF(TODAY()>D{r},"⚠️ OVERDUE","🔔 Upcoming"))')
            write_formula(8, f"=D{r}-14")
            write_formula(9, f"=D{r}-7")

            # Notes column (J)
            notes = payment.get("notes", "")
            notes_cell = ws.cell(row=excel_row, column=10, value=notes)
            if notes:
                notes_cell.fill = _FILL_GREEN
            notes_cell.font = _FONT_BLACK

        wb.save(self.filepath)

    def _remove_creditor(self, creditor_name: str) -> None:
        """Remove all rows belonging to creditor_name from the Excel file."""
        wb = openpyxl.load_workbook(self.filepath, data_only=False)
        ws = wb.active

        # Collect row indices (1-based) that belong to this creditor
        rows_to_delete = [
            row_idx
            for row_idx in range(1, ws.max_row + 1)
            if ws.cell(row=row_idx, column=1).value == creditor_name
        ]

        if not rows_to_delete:
            raise SettlementNotFoundError(f"Creditor '{creditor_name}' not found")

        # Expand deletion range to include adjacent blank separator rows
        first = rows_to_delete[0]
        last = rows_to_delete[-1]

        # Include blank row before the block (separator above), if any
        if first > 4 and ws.cell(row=first - 1, column=1).value is None:
            first -= 1

        # Include blank row after the block (separator below), if any
        if last < ws.max_row and ws.cell(row=last + 1, column=1).value is None:
            last += 1

        # Delete rows from bottom up to avoid index shifting
        for row_idx in range(last, first - 1, -1):
            ws.delete_rows(row_idx)

        wb.save(self.filepath)
