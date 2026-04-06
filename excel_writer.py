from __future__ import annotations

import calendar
from datetime import date, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font

# Cell fill / font colours matching the existing RememberMe template
_FILL_BLUE = PatternFill("solid", fgColor="FFDDEEFF")   # input columns A-E
_FILL_GREY = PatternFill("solid", fgColor="FFF2F2F2")   # formula columns F-I
_FILL_GREEN = PatternFill("solid", fgColor="FFE2EFDA")  # final-row notes col J
_FONT_BLUE = Font(color="FF0000FF")
_FONT_BLACK = Font(color="FF000000")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_payment_rows(
    creditor: str,
    total_amount: float,
    monthly_payment: float,
    first_due_date: date,
    payment_day_of_month: Optional[int] = None,
) -> list[dict]:
    """
    Generate a full payment schedule for one creditor.

    Returns a list of payment dicts compatible with reader.py's key contract:
      creditor, payment_num, amount, due_date, balance_before, balance_after (None),
      notes (extra — used by append_to_excel for the J column).
    """
    rows: list[dict] = []
    running_balance = float(total_amount)
    payment_num = 1
    due_date = first_due_date

    while running_balance > 0.0001:
        is_last = running_balance <= monthly_payment + 0.0001
        if is_last:
            # Compute final amount as total minus everything already scheduled
            # to avoid floating-point accumulation across many payments.
            amount = round(total_amount - sum(r["amount"] for r in rows), 2)
        else:
            amount = round(monthly_payment, 2)

        balance_before = round(running_balance, 2)
        running_balance = round(running_balance - amount, 10)

        rows.append({
            "creditor": creditor,
            "payment_num": payment_num,
            "amount": amount,
            "due_date": due_date,
            "balance_before": balance_before,
            "balance_after": None,   # Excel formula handles this
            "notes": "← Settlement complete!" if is_last else "",
        })

        if is_last:
            break

        payment_num += 1
        due_date = _next_due_date(due_date, payment_day_of_month)

    return rows


def append_to_excel(
    payments: list[dict],
    filepath: str = "RememberMe.xlsx",
) -> None:
    """
    Append a list of payment dicts to an existing RememberMe.xlsx file.

    Writes columns A-J with correct values and Excel formula strings.
    Preserves existing content and inserts one blank separator row before
    the new block.

    Raises FileNotFoundError if filepath does not exist.
    """
    import os
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"No such file: {filepath}")

    # data_only=False preserves existing formulas
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active

    # Find the first empty row after all existing data
    last_data_row = 3  # at minimum, 3 header rows exist
    for row_idx in range(ws.max_row, 3, -1):
        if ws.cell(row=row_idx, column=1).value is not None:
            last_data_row = row_idx
            break

    # Insert one blank separator row
    separator_row = last_data_row + 1
    first_data_row = separator_row + 1

    for i, payment in enumerate(payments):
        excel_row = first_data_row + i

        # --- Input columns (A-E): blue fill, blue font ---
        def write_input(col, value):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.fill = _FILL_BLUE
            cell.font = _FONT_BLUE

        write_input(1, payment["creditor"])
        write_input(2, payment["payment_num"])
        write_input(3, payment["amount"])
        # due_date: write as a Python date so openpyxl stores it as a date serial
        write_input(4, payment["due_date"])
        write_input(5, payment["balance_before"])

        # --- Formula columns (F-I): grey fill, black font ---
        def write_formula(col, formula):
            cell = ws.cell(row=excel_row, column=col, value=formula)
            cell.fill = _FILL_GREY
            cell.font = _FONT_BLACK

        r = excel_row
        write_formula(6, f"=E{r}-C{r}")
        write_formula(7, f'=IF(F{r}=0,"✅ PAID OFF",IF(TODAY()>D{r},"⚠️ OVERDUE","🔔 Upcoming"))')
        write_formula(8, f"=D{r}-14")
        write_formula(9, f"=D{r}-7")

        # --- Notes column (J) ---
        notes = payment.get("notes", "")
        notes_cell = ws.cell(row=excel_row, column=10, value=notes)
        if notes:
            notes_cell.fill = _FILL_GREEN
        notes_cell.font = _FONT_BLACK

    wb.save(filepath)


def remove_creditor(
    creditor_name: str,
    filepath: str = "RememberMe.xlsx",
) -> None:
    """
    Remove all rows belonging to creditor_name from the Excel file.

    Also removes any blank separator rows that become orphaned directly
    above or below the deleted block, leaving the file clean.

    Raises FileNotFoundError if filepath does not exist.
    Raises ValueError if creditor_name is not found in the file.
    """
    import os
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"No such file: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active

    # Collect row indices (1-based) that belong to this creditor
    rows_to_delete = [
        row_idx
        for row_idx in range(1, ws.max_row + 1)
        if ws.cell(row=row_idx, column=1).value == creditor_name
    ]

    if not rows_to_delete:
        raise ValueError(f"Creditor '{creditor_name}' not found in {filepath}")

    # Expand deletion range to include adjacent blank separator rows.
    # We include blank rows immediately before the block (if row > 3)
    # and immediately after the block.
    first = rows_to_delete[0]
    last = rows_to_delete[-1]

    # Include the blank row before the block (separator above), if any
    if first > 4 and ws.cell(row=first - 1, column=1).value is None:
        first -= 1

    # Include the blank row after the block (separator below), if any
    if last < ws.max_row and ws.cell(row=last + 1, column=1).value is None:
        last += 1

    # Delete rows from bottom up to avoid index shifting
    for row_idx in range(last, first - 1, -1):
        ws.delete_rows(row_idx)

    wb.save(filepath)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _next_due_date(current: date, payment_day_of_month: Optional[int]) -> date:
    """Advance to the next due date."""
    if not payment_day_of_month:
        return current + timedelta(days=28)

    # Advance one calendar month, then set the day (clamped to last day of month)
    month = current.month + 1
    year = current.year
    if month > 12:
        month = 1
        year += 1

    last_day = calendar.monthrange(year, month)[1]
    day = min(payment_day_of_month, last_day)
    return date(year, month, day)
